# content/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django_ratelimit.decorators import ratelimit
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import models
from django.core.cache import cache
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_headers

from .models import ContentItem, ContentItemPhoto, Group, Subcategory, Category, DisplayType, VideoCard
from django.core.paginator import Paginator
from .forms import BulkImportForm

import os
import zipfile
from django.conf import settings
from openpyxl import load_workbook


# ====================== Универсальная HTMX-функция ======================
def _htmx_or_full(request, partial_template, full_context):
    """Возвращает partial при HTMX-запросе, иначе — полный контекст"""
    if request.htmx:
        return render(request, partial_template, full_context)
    return full_context


@cache_page(60 * 2)
def home(request):
    total_prompts = cache.get_or_set('home_total_prompts', lambda: ContentItem.objects.count(), 300)
    photos = cache.get_or_set(
        'home_photos',
        lambda: list(ContentItemPhoto.objects.filter(
            content_item__category__slug__in=['photo-sessions', 'man-photo-sessions']
        ).select_related('content_item').order_by('-created_at')[:15]),
        300
    )
    photo_urls = [p.photo.url for p in photos if p.photo]
    return render(request, 'content/home.html', {'total_prompts': total_prompts, 'photo_urls': photo_urls})


@require_POST
@ratelimit(key='user', rate='10/m', block=True)
def copy_content(request, pk):
    try:
        item = ContentItem.objects.get(pk=pk)
        return JsonResponse({'success': True, 'text': item.full_text})
    except ContentItem.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Запись не найдена'}, status=404)


@cache_page(60 * 5)
@vary_on_headers('HX-Request')
def content_list(request):
    qs = ContentItem.objects.select_related('category', 'subcategory__category', 'group')\
                            .prefetch_related('photos').order_by('-created_at')
    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'items': page_obj.object_list,
        'page_obj': page_obj,
        'is_last_page': not page_obj.has_next(),
    }

    if request.htmx:
        return render(request, 'content/_card_list.html', context)

    top_categories = Category.objects.filter(order__gt=0).order_by('order')
    categories = Subcategory.objects.select_related('category').order_by('category__order', 'name')

    return render(request, 'content/list.html', {
        **context,
        'top_categories': top_categories,
        'categories': categories,
    })


@cache_page(60 * 5)
@vary_on_headers('HX-Request')
def category_detail(request, slug):
    top_category = get_object_or_404(Category, slug=slug)
    items = ContentItem.objects.filter(
        models.Q(category=top_category) | models.Q(subcategory__category=top_category)
    ).select_related('category', 'subcategory__category', 'group').prefetch_related('photos').distinct().order_by('-created_at')

    paginator = Paginator(items, 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'category': top_category,
        'top_category': top_category,
        'is_top_category': True,
        'items': page_obj.object_list,
        'page_obj': page_obj,
        'is_last_page': not page_obj.has_next(),
    }

    if request.htmx:
        return render(request, 'content/_card_list.html', context)

    top_categories = Category.objects.filter(order__gt=0).order_by('order')
    subcategories = Subcategory.objects.filter(category=top_category).order_by('name')

    return render(request, 'content/category_detail.html', {
        **context,
        'categories': subcategories,
        'top_categories': top_categories,
    })


@cache_page(60 * 5)
@vary_on_headers('HX-Request')
def subcategory_detail(request, category_slug, subcategory_slug):
    top_category = get_object_or_404(Category, slug=category_slug)
    subcategory = get_object_or_404(Subcategory, slug=subcategory_slug, category=top_category)
    items = ContentItem.objects.filter(subcategory=subcategory)\
        .select_related('category', 'group').prefetch_related('photos').order_by('-created_at')

    paginator = Paginator(items, 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'category': subcategory,
        'top_category': top_category,
        'is_top_category': False,
        'items': page_obj.object_list,
        'page_obj': page_obj,
        'is_last_page': not page_obj.has_next(),
    }

    if request.htmx:
        return render(request, 'content/_card_list.html', context)

    top_categories = Category.objects.filter(order__gt=0).order_by('order')
    subcategories = Subcategory.objects.filter(category=top_category).order_by('name')

    return render(request, 'content/category_detail.html', {
        **context,
        'categories': subcategories,
        'top_categories': top_categories,
    })


@cache_page(60 * 5)
def group_detail(request, slug):
    group = get_object_or_404(Group, slug=slug)
    items = ContentItem.objects.filter(group=group)\
        .select_related('category', 'subcategory__category').prefetch_related('photos').order_by('-created_at')

    top_categories = Category.objects.filter(order__gt=0).order_by('order')
    categories = Subcategory.objects.select_related('category').order_by('category__order', 'name')

    return render(request, 'content/group_detail.html', {
        'group': group,
        'items': items,
        'categories': categories,
        'top_categories': top_categories,
    })


# ====================== ВИДЕО ======================
@cache_page(60 * 5)
def video_list(request):
    qs = VideoCard.objects.filter(is_active=True).prefetch_related('references').order_by('-created_at')

    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'video_cards': page_obj.object_list,
        'page_obj': page_obj,
        'is_last_page': not page_obj.has_next(),
        'is_video_page': True,
    }

    if request.htmx:
        return render(request, 'content/_video_card_list.html', context)

    return render(request, 'content/video_list.html', context)

@require_POST
@ratelimit(key='user_or_ip', rate='10/m', block=True)
def copy_video_card(request, pk):
    try:
        card = VideoCard.objects.get(pk=pk, is_active=True)

        return JsonResponse({
            'success': True,
            'text': card.copy_text
        })
    except VideoCard.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)  

# ====================== Bulk Import ======================
@staff_member_required
def bulk_import_view(request):
    """Массовый импорт с поддержкой display_type"""
    if request.method == 'POST':
        form = BulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            photos_zip = request.FILES.get('photos_zip')

            if photos_zip:
                extract_path = os.path.join(settings.MEDIA_ROOT, 'photos')
                os.makedirs(extract_path, exist_ok=True)
                safe_root = os.path.realpath(extract_path)
                MAX_TOTAL_BYTES = 2 * 1024 * 1024 * 1024  # 2 GB защита от zip bomb
                MAX_FILES = 5000

                try:
                    with zipfile.ZipFile(photos_zip, 'r') as zip_ref:
                        members = zip_ref.infolist()
                        if len(members) > MAX_FILES:
                            raise ValueError(f"Слишком много файлов в архиве: {len(members)}")
                        total = 0
                        for m in members:
                            total += m.file_size
                            if total > MAX_TOTAL_BYTES:
                                raise ValueError("Распакованный размер архива превышает лимит")
                            dest = os.path.realpath(os.path.join(safe_root, m.filename))
                            if not dest.startswith(safe_root + os.sep) and dest != safe_root:
                                raise ValueError(f"Подозрительный путь в архиве: {m.filename}")
                        zip_ref.extractall(extract_path)
                    messages.info(request, "✅ Фотографии распакованы")
                except (zipfile.BadZipFile, ValueError) as e:
                    messages.error(request, f"❌ Ошибка распаковки ZIP: {e}")
                    return redirect('admin:bulk_import')

            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active

            created = 0
            errors = []

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                values = [str(cell).strip() if cell is not None else '' for cell in row]

                filename = values[0] if len(values) > 0 else ''
                category_name = values[1] if len(values) > 1 else ''
                subcategory_name = values[2] if len(values) > 2 else ''
                group_name = values[3] if len(values) > 3 else ''
                prompt = values[4] if len(values) > 4 else ''
                photos_str = values[5] if len(values) > 5 else ''

                if not filename or not category_name or not prompt:
                    errors.append(f"Строка {row_num}: filename, category и prompt обязательны")
                    continue

                try:
                    category = Category.objects.get(name__iexact=category_name)
                except Category.DoesNotExist:
                    errors.append(f"Строка {row_num}: категория «{category_name}» не найдена")
                    continue

                subcategory = Subcategory.objects.filter(
                    category=category, name__iexact=subcategory_name
                ).first() if subcategory_name else None

                group = Group.objects.filter(name__iexact=group_name).first() if group_name else None

                # === Главное изменение ===
                item = ContentItem.objects.create(
                    category=category,
                    subcategory=subcategory,
                    group=group,
                    full_text=prompt,
                    display_type=category.display_type,   # ← берём default
                )

                # Фото
                photo_list = [f.strip() for f in photos_str.split(',') if f.strip()] if photos_str else [filename]
                for idx, photo_file in enumerate(photo_list):
                    photo_path = os.path.join('photos', photo_file)
                    full_path = os.path.join(settings.MEDIA_ROOT, photo_path)
                    if os.path.exists(full_path):
                        ContentItemPhoto.objects.create(
                            content_item=item,
                            photo=photo_path,
                            order=idx
                        )

                created += 1

            if created:
                messages.success(request, f'✅ Успешно создано {created} карточек.')
            if errors:
                for err in errors[:15]:
                    messages.error(request, err)

            return redirect('admin:content_contentitem_changelist')

    else:
        form = BulkImportForm()

    return render(request, 'content/bulk_import.html', {'form': form})

